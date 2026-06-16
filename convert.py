"""
convert.py

Unified entry point: accepts .doc, .docx, or .pdf and routes to the
appropriate pipeline (docx_parsing or pdf_parsing).

Single-file usage:
    python convert.py input/file.docx
    python convert.py input/file.doc
    python convert.py input/file.pdf
    python convert.py input/file.pdf --ocr-engine surya
    python convert.py input/file.docx -f gfm
    python convert.py input/file.docx output/custom.md
    python convert.py input/file.pdf --flow pymupdf4llm        # force-OCR via pymupdf4llm

Folder (batch) usage:
    python convert.py input/folder/
    python convert.py input/folder/ output/folder/
    python convert.py input/folder/ --ocr-engine tesseract
    python convert.py input/folder/ --flow pymupdf4llm         # pymupdf4llm flow for all PDFs

    Already-converted files are skipped based on (filename, size).
    Tracking state is persisted in <output_dir>/.convert_state.json.
    An Excel summary is written to <output_dir>/conversion_summary_YYYYMMDD_HHMMSS.xlsx.

Library usage:
    from convert import convert, batch_convert
    md = convert("input/file.pdf")
    md = convert("input/file.pdf", flow="pymupdf4llm")
    excel = batch_convert("input/folder/")
    excel = batch_convert("input/folder/", flow="pymupdf4llm")
"""

import json
import logging
import os
import sys
import time
from datetime import datetime

SUPPORTED_EXTENSIONS = {".doc", ".docx", ".pdf"}

log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# Single-file conversion
# ══════════════════════════════════════════════════════════════════════════════

def convert(
    input_path: str,
    output_md: str = None,
    md_format: str = "markdown",
    ocr_engine: str = "surya",
    ocr_equations: bool = False,
    dpi: int = 300,
    flow: str = "default",
) -> str:
    """
    Detect file type and route to the correct conversion pipeline.

    Parameters
    ----------
    input_path : str
        Path to a .doc, .docx, or .pdf file.
    output_md : str, optional
        Where to write the Markdown output. Defaults to output/<stem>.md.
    md_format : str
        Pandoc output format (only used for doc/docx). Default: "markdown".
    ocr_engine : str
        OCR engine for scanned PDF pages: "surya" (default), "doctr", or "tesseract".
    ocr_equations : bool
        If True, OCR equation images in .doc files (via pix2tex) and replace them
        with LaTeX text. Requires: pip install pix2tex. Default: False.
    dpi : int
        Resolution for OCR rendering (only used for PDF). Default: 300.
    flow : str
        PDF conversion flow: "default" (layered ins/del detection) or
        "pymupdf4llm" (force-OCR via pymupdf4llm, no ins/del tags).
        Ignored for doc/docx. Default: "default".

    Returns
    -------
    str
        The generated Markdown string.
    """
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"File not found: {input_path}")

    ext = os.path.splitext(input_path)[1].lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type '{ext}'. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    if output_md is None:
        stem = os.path.splitext(os.path.basename(input_path))[0]
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(output_dir, exist_ok=True)
        output_md = os.path.join(output_dir, stem + ".md")

    if ext in (".doc", ".docx"):
        from docx_parsing import convert as docx_convert
        return docx_convert(
            input_path, output_md=output_md, md_format=md_format,
            ocr_equations=ocr_equations,
        )

    # ext == ".pdf"
    from pdf_parsing import convert as pdf_convert
    return pdf_convert(input_path, output_md=output_md, ocr_engine=ocr_engine, dpi=dpi, flow=flow)


# ══════════════════════════════════════════════════════════════════════════════
# Batch / folder conversion
# ══════════════════════════════════════════════════════════════════════════════

_STATE_FILE = ".convert_state.json"


def _load_state(output_dir: str) -> dict:
    path = os.path.join(output_dir, _STATE_FILE)
    if os.path.isfile(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {"processed": {}}


def _save_state(state: dict, output_dir: str) -> None:
    path = os.path.join(output_dir, _STATE_FILE)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def _file_key(fname: str, fsize: int) -> str:
    """Stable dict key: filename + size identifies a unique source file."""
    return f"{fname}:{fsize}"


def _resolve_output_stem(stem: str, used_stems: set) -> str:
    """Return stem (possibly suffixed) that is not already in used_stems."""
    candidate = stem
    counter = 1
    while candidate in used_stems:
        candidate = f"{stem}_{counter}"
        counter += 1
    return candidate


def _write_excel(
    results: list,
    run_start: datetime,
    input_dir: str,
    output_dir: str,
    ocr_engine: str,
    md_format: str,
    total_elapsed: float,
) -> str:
    """Write Excel summary and return its path."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ts = run_start.strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"conversion_summary_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ── Detail sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "File Details"

    headers = [
        "Input File",
        "Input Path",
        "Input Size (bytes)",
        "Output File",
        "Output Path",
        "Method",
        "OCR Engine",
        "Status",
        "Conversion Start Time",
        "Elapsed (s)",
        "Error",
    ]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    fill_success = PatternFill("solid", fgColor="E2EFDA")   # light green
    fill_skipped = PatternFill("solid", fgColor="FFF2CC")   # light yellow
    fill_failed  = PatternFill("solid", fgColor="FCE4D6")   # light red

    for row_idx, r in enumerate(results, 2):
        values = [
            r["input_file"],
            r["input_path"],
            r["input_size_bytes"],
            r["output_file"],
            r["output_path"],
            r["method"],
            r["ocr_engine"],
            r["status"],
            r["start_time"],
            r["elapsed_seconds"],
            r["error"],
        ]
        status = r["status"]
        row_fill = (
            fill_success if status == "success"
            else fill_skipped if status.startswith("skipped")
            else fill_failed
        )
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill

    # Auto-size columns (cap at 60)
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 60)

    # ── Run summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Run Summary")
    successes = sum(1 for r in results if r["status"] == "success")
    failures  = sum(1 for r in results if r["status"] == "failed")
    skips     = sum(1 for r in results if r["status"].startswith("skipped"))

    summary_rows = [
        ("Run Started",              run_start.strftime("%Y-%m-%d %H:%M:%S")),
        ("Input Directory",          input_dir),
        ("Output Directory",         output_dir),
        ("Total Files Found",        len(results)),
        ("Newly Converted (Success)", successes),
        ("Failed",                   failures),
        ("Skipped (Already Done)",   skips),
        ("Total Elapsed (s)",        round(total_elapsed, 2)),
        ("OCR Engine",               ocr_engine),
        ("Markdown Format",          md_format),
    ]

    hdr_fill2 = PatternFill("solid", fgColor="2E75B6")
    hdr_font2 = Font(bold=True, color="FFFFFF")
    ws2.cell(row=1, column=1, value="Metric").fill = hdr_fill2
    ws2.cell(row=1, column=1).font = hdr_font2
    ws2.cell(row=1, column=2, value="Value").fill = hdr_fill2
    ws2.cell(row=1, column=2).font = hdr_font2

    for row_idx, (k, v) in enumerate(summary_rows, 2):
        ws2.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=v)

    for col in ws2.columns:
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 60)

    wb.save(excel_path)
    return excel_path


def batch_convert(
    input_dir: str,
    output_dir: str = None,
    md_format: str = "markdown",
    ocr_engine: str = "surya",
    ocr_equations: bool = False,
    dpi: int = 300,
    flow: str = "default",
) -> str:
    """
    Convert all supported files in input_dir to Markdown, skipping files
    that have already been converted (matched by filename + file size).

    Tracking state persists in <output_dir>/.convert_state.json so skips
    work correctly across separate runs.

    Parameters
    ----------
    input_dir     : str  — directory containing .doc / .docx / .pdf files
    output_dir    : str  — where to write .md files (default: input_dir/output)
    md_format     : str  — pandoc format for doc/docx (default: "markdown")
    ocr_engine    : str  — OCR engine for scanned PDFs (default: "surya")
    ocr_equations : bool — OCR equation images in .doc files via pix2tex (default: False)
    dpi           : int  — DPI for OCR rendering (default: 300)
    flow          : str  — PDF flow: "default" or "pymupdf4llm" (default: "default")

    Returns
    -------
    str — path to the Excel summary file that was written
    """
    from tqdm import tqdm

    input_dir = os.path.abspath(input_dir)
    if not os.path.isdir(input_dir):
        raise NotADirectoryError(f"Not a directory: {input_dir}")

    if output_dir is None:
        output_dir = os.path.join(input_dir, "output")
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    # ── Scan input directory ───────────────────────────────────────────────────
    all_files: list[tuple[str, str, int]] = []  # (fname, fpath, fsize)
    for fname in sorted(os.listdir(input_dir)):
        ext = os.path.splitext(fname)[1].lower()
        if ext in SUPPORTED_EXTENSIONS:
            fpath = os.path.join(input_dir, fname)
            fsize = os.path.getsize(fpath)
            all_files.append((fname, fpath, fsize))

    print(f"Found {len(all_files)} supported file(s) in {input_dir}")

    # ── Load persistent tracking state ────────────────────────────────────────
    state = _load_state(output_dir)

    # ── Partition: skip vs. process ───────────────────────────────────────────
    to_process: list[tuple[str, str, int]] = []
    skipped_entries: list[tuple[str, str, int, dict]] = []

    for fname, fpath, fsize in all_files:
        key = _file_key(fname, fsize)
        if key in state["processed"]:
            skipped_entries.append((fname, fpath, fsize, state["processed"][key]))
        else:
            to_process.append((fname, fpath, fsize))

    print(f"  Already converted : {len(skipped_entries)}")
    print(f"  To process        : {len(to_process)}")

    # ── Build set of already-used output stems ────────────────────────────────
    used_stems: set[str] = set()
    for entry in state["processed"].values():
        out = entry.get("output_file", "")
        if out:
            used_stems.add(os.path.splitext(os.path.basename(out))[0])
    for fname in os.listdir(output_dir):
        if fname.endswith(".md"):
            used_stems.add(os.path.splitext(fname)[0])

    # ── Build results list (skipped rows first) ───────────────────────────────
    results: list[dict] = []

    for fname, fpath, fsize, entry in skipped_entries:
        out_path = entry.get("output_file", "")
        results.append({
            "input_file":        fname,
            "input_path":        fpath,
            "input_size_bytes":  fsize,
            "output_file":       os.path.basename(out_path),
            "output_path":       out_path,
            "method":            entry.get("method", ""),
            "ocr_engine":        entry.get("ocr_engine", ""),
            "status":            "skipped (already converted)",
            "start_time":        entry.get("timestamp", ""),
            "elapsed_seconds":   entry.get("elapsed_seconds", ""),
            "error":             "",
        })

    # ── Process new files ─────────────────────────────────────────────────────
    run_start   = datetime.now()
    batch_t0    = time.time()

    bar = tqdm(to_process, desc="Converting files", unit="file", dynamic_ncols=True)
    for fname, fpath, fsize in bar:
        bar.set_postfix_str(fname)

        ext    = os.path.splitext(fname)[1].lower()
        stem   = os.path.splitext(fname)[0]
        method = "docx" if ext in (".doc", ".docx") else "pdf"

        out_stem  = _resolve_output_stem(stem, used_stems)
        used_stems.add(out_stem)
        output_md = os.path.join(output_dir, out_stem + ".md")

        start_dt = datetime.now()
        t_start  = time.time()
        status   = "success"
        error    = ""

        try:
            convert(
                fpath,
                output_md=output_md,
                md_format=md_format,
                ocr_engine=ocr_engine,
                ocr_equations=ocr_equations,
                dpi=dpi,
                flow=flow,
            )
        except Exception as exc:
            status    = "failed"
            error     = str(exc)
            output_md = ""
            log.error("Failed to convert %s: %s", fname, exc)

        elapsed = round(time.time() - t_start, 2)

        # Persist to state
        key = _file_key(fname, fsize)
        state["processed"][key] = {
            "input_path":      fpath,
            "output_file":     output_md,
            "method":          method,
            "ocr_engine":      ocr_engine if method == "pdf" else "",
            "status":          status,
            "timestamp":       start_dt.isoformat(timespec="seconds"),
            "elapsed_seconds": elapsed,
        }
        _save_state(state, output_dir)  # save after each file so progress survives interrupts

        results.append({
            "input_file":       fname,
            "input_path":       fpath,
            "input_size_bytes": fsize,
            "output_file":      os.path.basename(output_md) if output_md else "",
            "output_path":      output_md,
            "method":           method,
            "ocr_engine":       ocr_engine if method == "pdf" else "",
            "status":           status,
            "start_time":       start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "elapsed_seconds":  elapsed,
            "error":            error,
        })

    total_elapsed = time.time() - batch_t0

    # ── Write Excel summary ────────────────────────────────────────────────────
    excel_path = _write_excel(
        results=results,
        run_start=run_start,
        input_dir=input_dir,
        output_dir=output_dir,
        ocr_engine=ocr_engine,
        md_format=md_format,
        total_elapsed=total_elapsed,
    )

    # ── Console run summary ────────────────────────────────────────────────────
    successes = sum(1 for r in results if r["status"] == "success")
    failures  = sum(1 for r in results if r["status"] == "failed")
    skips     = sum(1 for r in results if r["status"].startswith("skipped"))

    print(f"\n{'=' * 60}")
    print("Run Summary")
    print(f"{'=' * 60}")
    print(f"  Input dir   : {input_dir}")
    print(f"  Output dir  : {output_dir}")
    print(f"  Total files : {len(results)}")
    print(f"  Converted   : {successes}")
    print(f"  Skipped     : {skips}")
    print(f"  Failed      : {failures}")
    print(f"  Elapsed     : {total_elapsed:.1f}s")
    print(f"  Excel report: {excel_path}")
    print(f"{'=' * 60}\n")

    return excel_path


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Convert .doc/.docx/.pdf with track changes to Markdown with <ins>/<del> tags.\n"
            "Pass a file for single-file mode or a directory for batch mode."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input",
        help="Path to a .doc/.docx/.pdf file, or a directory for batch mode.",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=None,
        help=(
            "Output path: .md file (single-file mode) or directory (batch mode). "
            "Defaults to output/<stem>.md or <input_dir>/output/ respectively."
        ),
    )
    parser.add_argument(
        "-f", "--format",
        default="markdown",
        dest="md_format",
        help="Pandoc output format for doc/docx: markdown (default), gfm, etc.",
    )
    parser.add_argument(
        "--ocr-engine",
        choices=["surya", "doctr", "tesseract"],
        default="surya",
        help="OCR engine for scanned PDF pages (default: surya).",
    )
    parser.add_argument(
        "--ocr-equations",
        action="store_true",
        default=False,
        help=(
            "OCR equation images in .doc files (equations that LibreOffice rendered as "
            "PNG instead of preserving as math) and replace them with LaTeX text. "
            "Requires: pip install pix2tex"
        ),
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=300,
        help="DPI for OCR page rendering (default: 300).",
    )
    parser.add_argument(
        "--flow",
        choices=["default", "pymupdf4llm"],
        default="default",
        help="PDF conversion flow: 'default' (layered ins/del detection) or "
             "'pymupdf4llm' (force-OCR via pymupdf4llm, no ins/del tags). "
             "Ignored for doc/docx. Default: default.",
    )

    args = parser.parse_args()

    if os.path.isdir(args.input):
        # Batch mode
        excel = batch_convert(
            input_dir=args.input,
            output_dir=args.output,
            md_format=args.md_format,
            ocr_engine=args.ocr_engine,
            ocr_equations=args.ocr_equations,
            dpi=args.dpi,
            flow=args.flow,
        )
        print(f"Excel summary written to: {excel}")
    else:
        # Single-file mode
        md = convert(
            args.input,
            output_md=args.output,
            md_format=args.md_format,
            ocr_engine=args.ocr_engine,
            ocr_equations=args.ocr_equations,
            dpi=args.dpi,
            flow=args.flow,
        )
        print(md)
